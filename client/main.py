import PySimpleGUI as sg
import sqlite3
import hashlib
import datetime
import os
import sys
import time
import math as m
import pandas as pd
from openpyxl import load_workbook


conn = sqlite3.connect("database.db")
cursor = conn.cursor()
roles = [
    "Administrator",
    "Kladovshik",
    "Injener",
    "Testirovshik",
    "Rukovoditel proizvodstva",
    "Rukovoditel otdela zakupok",
    "Menedjer po zakupkam",
    "Rukovoditel otdela prodaj",
    "Menedjer po prodajam",
]
permitRoles = [
    "Administrator",
    "Kladovshik",
    "Injener",
    "Rukovoditel proizvodstva",
    "Rukovoditel otdela zakupok",
    "Menedjer po zakupkam",
]


def excelToDB(path):
    """
    Reads an Excel file from the specified path and inserts its data into a database table.

    Parameters:
        path (str): The path to the Excel file.

    Returns:
        bool: True if the data was successfully inserted into the database, False otherwise.
    """
    # TODO: Read an Excel file and insert its data into a database table

    wb = load_workbook(filename=path)
    sheet = wb.active
    cc = []
    bb = []
    skip = True
    for row in sheet.iter_rows(values_only=True):
        for i in range(5):
            bb.append(str(list(row)[i]).replace("\n", " "))
        cc.append(bb)
        print(bb)
        if skip:
            bb = []
            continue
            skip = False

        insert_query = "INSERT INTO components (name, article, type, weight, out_date) VALUES (?, ?, ?, ?, ?)"
        cursor.execute(insert_query, bb)
        bb = []
    conn.commit()


def extractDBtoExcel(database, path):
    """
    Extracts data from a database table to an Excel file.

    Parameters:
        path (str): The path to the Excel file.

    Returns:
        bool: True if the data was successfully extracted to the Excel file, False otherwise.
    """
    # TODO: Extract data from a database table to an Excel file
    try:
        df = pd.read_sql(f"SELECT * FROM {database}", conn)
        df.to_excel(path)
        return True
    except:
        return False


def sanitizeInput(input):
    """
    A function that sanitizes the input by removing any forbidden characters.

    Parameters:
    input (str): The input string to be sanitized.

    Returns:
    str: The sanitized input string.
    """
    ##TODO: Sanitize input
    forbiddenChars = [
        "\\",
        "/",
        ":",
        "*",
        "?",
        '"',
        "<",
        ">",
        "|",
        " ",
        "(",
        ")",
        "$",
        "%",
        "&",
        "卐",
        "-",
        "_",
        "+",
        "=",
        "[",
        "]",
        "{",
        "}",
        ";",
        "!",
        "@",
        "#",
        "^",
        "`",
        "~",
        ",",
        ".",
    ]

    for char in input:
        if char in forbiddenChars:
            input = input.replace(char, "")
    return input


def initializeDatabase():
    """
    Initialize the database.
    """
    ##TODO: Initialize the database
    cursor = conn.cursor()
    cursor.execute(
        "CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY AUTOINCREMENT, name VARCHAR(255), password VARCHAR(255), role VARCHAR(255) NULL, last_login DATETIME, last_logout DATETIME, last_login_device VARCHAR(255) NULL)"
    )
    cursor.execute(
        "CREATE TABLE IF NOT EXISTS components (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, article VARCHAR(255), type VARCHAR(255), weight REAL, out_date DATETIME)"
    )

    conn.commit()
    return True


def registration(name, password):
    """
    Register a new user in the system.

    Parameters:
    name (str): The name of the user to be registered.
    password (str): The password of the user to be registered.

    Returns:
    bool: True if the user was successfully registered, False if the user already exists.
    """
    ##TODO: Register a new user
    cursor.execute("SELECT * FROM users WHERE name = ?", (name,))
    if cursor.fetchone():
        print("User already exists")
        return False
    else:
        hashed_password = hash_password(password)
        cursor.execute(
            "INSERT INTO users (name, password) VALUES (?, ?)", (name, hashed_password)
        )
        conn.commit()
        return True


def hash_password(password):
    """
    Hash a password using sha256 algorithm.

    Parameters:
    password (str): The password to be hashed.

    Returns:
    str: The hashed password.
    """
    ##TODO: Hash a password
    return hashlib.sha256(password.encode()).hexdigest()


def checkRole(user_id):
    """
    Check role of a user
    :param user_id: The ID of the user to check the role for
    :return: The role of the user
    """
    ##TODO: Check role of a user
    cursor.execute("SELECT role FROM users WHERE id = ?", (user_id,))
    role = cursor.fetchone()[0]
    return role


def login(name, password, device):
    """
    Login a user.

    Parameters:
    - name (str): The name of the user.
    - password (str): The password of the user.
    - device (str): The device used for login.

    Returns:
    - user (dict): The user information if login is successful, False otherwise.
    """
    ##TODO: Login a user
    hashed_password = hash_password(password)
    cursor.execute(
        "SELECT * FROM users WHERE name = ? AND password = ?", (name, hashed_password)
    )
    user = cursor.fetchone()
    if user:
        cursor.execute(
            "UPDATE users SET last_login = ?, last_login_device = ? WHERE name = ?",
            (datetime.datetime.now(), device, name),
        )
        conn.commit()
        return user
    else:
        print("Wrong name or password")
        return None


def logout(user_id):
    """
    Logout a user by updating the last_logout timestamp in the users table for the given user_id.

    Args:
        user_id: The unique identifier of the user to be logged out.

    Returns:
        bool: True if the user is successfully logged out, False otherwise.
    """
    ##TODO: Logout a user
    cursor.execute("UPDATE users SET last_logout = DATETIME() WHERE id = ?", (user_id,))
    conn.commit()
    return True


def createComponent(user_id, name, article, type, weight, out_date):
    ##TODO: Create a new component
    if checkRole(user_id) in permitRoles:
        cursor.execute(
            "INSERT INTO components (name, article, type, weight, out_date) VALUES (?, ?, ?, ?, ?)",
            (name, article, type, weight, out_date),
        )
        conn.commit()
        return True
    else:
        print("Your role is not permitted to create components")
        return False


def changeName(admin_id, target_user_id, new_name):
    """
    Change name of a user
    Parameters:
    - admin_id: int, the ID of the administrator performing the change
    - target_user_id: int, the ID of the user whose name will be changed
    - new_name: str, the new name to assign to the user
    Returns:
    - bool, True if the name was successfully changed, False otherwise
    """
    # TODO: Change name of a user
    if checkRole(admin_id) == "Administrator":
        cursor.execute(
            "UPDATE users SET name = ? WHERE id = ?", (new_name, target_user_id)
        )
        conn.commit()
        return True
    else:
        print("Only administrators can change names")
        return False


def changeRole(admin_id, target_user_id, role):
    """
    Change role of a user.

    Parameters:
    - admin_id (int): The user ID of the administrator performing the role change.
    - target_user_id (int): The user ID of the user whose role is being changed.
    - role (str): The new role to assign to the user.

    Returns:
    - bool: True if the role change was successful, False otherwise.
    """
    # TODO: Change role of a user
    if checkRole(admin_id) == "Administrator":
        cursor.execute("UPDATE users SET role = ? WHERE id = ?", (role, target_user_id))
        conn.commit()
        return True
    else:
        print("Only Administrators can change roles")
        return False


def changePassword(admin_id, target_user_id, new_password):
    """
    Change password of a user

    Args:
        admin_id: The ID of the administrator performing the password change
        target_user_id: The ID of the user whose password is being changed
        new_password: The new password for the user

    Returns:
        bool: True if the password was successfully changed, False otherwise
    """
    # TODO: Change password of a user
    if checkRole(admin_id) == "Administrator":
        cursor.execute(
            "UPDATE users SET password = ? WHERE id = ?",
            (hash_password(new_password), target_user_id),
        )
        conn.commit()
        return True
    else:
        print("Only administrators can change passwords")
        return False


def deleteUser(admin_id, target_user_id):
    """
    Delete a user if the admin is an administrator and the target user is not an administrator.

    Parameters:
    - admin_id: the id of the admin user
    - target_user_id: the id of the user to be deleted

    Returns:
    - True if the user is successfully deleted, False otherwise
    """
    # TODO: Delete a user
    if (
        checkRole(admin_id) == "Administrator"
        and checkRole(target_user_id) != "Administrator"
    ):
        cursor.execute("DELETE FROM users WHERE id = ?", (target_user_id,))
        conn.commit()
        return True
    else:
        print("Only administrators can delete users")
        return False


def getUsers(admin_id):
    """
    Get all users using the given admin ID and return the list of users.
    Parameters:
    - admin_id: string, the ID of the administrator
    Returns:
    - list of lists, each inner list representing a user's data
    - False if the admin ID does not have the 'Administrator' role
    """
    # TODO: Get all users
    if checkRole(admin_id) == "Administrator":
        cursor.execute("SELECT * FROM users")
        result = [list(i) for i in cursor.fetchall()]
        return result
    else:
        print("Only administrators can get users")
        return False


def getComponents(user_id):
    """
    Get all components based on the user ID if the user's role is permitted, otherwise return False.

    Parameters:
    user_id (int): The ID of the user to check permissions for.

    Returns:
    list or bool: A list of components if the user's role is permitted, False otherwise.
    """
    # TODO: Get all components
    if checkRole(user_id) in permitRoles:
        cursor.execute("SELECT * FROM components")
        result = [list(i) for i in cursor.fetchall()]
        return result
    else:
        print("Your role is not permitted to get components")
        return False


def changeComponentName(user_id, component_id, new_name):
    """
    Change name of a component
    user_id: int, the user's ID
    component_id: int, the component's ID
    new_name: str, the new name for the component
    Returns True if the component name was successfully changed, False otherwise
    """
    # TODO: Change name of a component
    if checkRole(user_id) in permitRoles:
        cursor.execute(
            "UPDATE components SET name = ? WHERE id = ?", (new_name, component_id)
        )
        conn.commit()
        return True
    else:
        print("Your role is not permitted to change components")
        return False


def changeComponentArticle(user_id, component_id, new_article):
    """
    Change article of a component.

    Parameters:
    - user_id: int, the user ID
    - component_id: int, the component ID
    - new_article: str, the new article

    Returns:
    - bool: True if the article was successfully changed, False otherwise
    """
    # TODO: Change article of a component
    if checkRole(user_id) in permitRoles:
        cursor.execute(
            "UPDATE components SET article = ? WHERE id = ?",
            (new_article, component_id),
        )
        conn.commit()
        return True
    else:
        print("Your role is not permitted to change components")
        return False


def changeComponentType(user_id, component_id, new_type):
    """
    Change type of a component.

    Parameters:
    user_id (int): The user ID.
    component_id (int): The component ID.
    new_type (str): The new type to assign to the component.

    Returns:
    bool: True if the type of the component was successfully changed, False otherwise.
    """
    # TODO: Change type of a component
    if checkRole(user_id) in permitRoles:
        cursor.execute(
            "UPDATE components SET type = ? WHERE id = ?", (new_type, component_id)
        )
        conn.commit()
        return True
    else:
        print("Your role is not permitted to change components")
        return False


def changeComponentWeight(user_id, component_id, new_weight):
    """
    Change weight of a component
    Parameters:
    - user_id: the id of the user making the change
    - component_id: the id of the component to change
    - new_weight: the new weight to set for the component
    Returns:
    - True if the weight change is successful, False otherwise
    """
    # TODO: Change weight of a component
    if checkRole(user_id) in permitRoles:
        cursor.execute(
            "UPDATE components SET weight = ? WHERE id = ?", (new_weight, component_id)
        )
        conn.commit()
        return True
    else:
        print("Your role is not permitted to change components")
        return False


def changeComponentProductionDate(user_id, component_id, new_production_date):
    """
    Change production date of a component.

    Parameters:
    - user_id (int): The user ID attempting to change the production date.
    - component_id (int): The ID of the component whose production date is to be changed.
    - new_production_date (str): The new production date to be set for the component.

    Returns:
    - bool: True if the production date was successfully updated, False otherwise.
    """
    # TODO: Change production date of a component
    if checkRole(user_id) in permitRoles:
        cursor.execute(
            "UPDATE components SET production_date = ? WHERE id = ?",
            (new_production_date, component_id),
        )
        conn.commit()
        return True
    else:
        print("Your role is not permitted to change components")
        return False


def getCerteinComponents(user_id, component_type):
    """
    Get certain components based on the user ID and component type.
    :param user_id: the ID of the user
    :param component_type: the type of component to retrieve
    :return: a list of components if the user's role permits, otherwise False
    """
    # TODO: Get certain components
    if checkRole(user_id) in permitRoles:
        cursor.execute("SELECT * FROM components WHERE type = ?", (component_type,))
        result = [list(i) for i in cursor.fetchall()]
        return result
    else:
        print("Your role is not permitted to get certain components")
        return False


def authWindow():
    c = 0
    sg.theme("DarkAmber")
    status = None
    layout = [
        [sg.Text("", key="-STATE-")],
        [sg.Text("Enter login")],
        [sg.InputText()],
        [sg.Text("Enter password")],
        [sg.InputText()],
        [sg.Button("Login"), sg.Button("Register")],
    ]
    window = sg.Window("Autorization window", layout)
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        if event == "Register":
            registration(sanitizeInput(values[0]), sanitizeInput(values[1]))
        if event == "Login":
            status = login(
                sanitizeInput(values[0]),
                sanitizeInput(values[1]),
                os.environ.get("USERNAME"),
            )
            if status == None:
                if c <= 2:
                    window["-STATE-"].Update("Wrong login or password")
                    c += 1
                    timer = time.time()
                else:
                    window["-STATE-"].Update(
                        f"Wrong login or password, {m.ceil(300 - (time.time() - timer))} seconds left"
                    )
                    if time.time() - timer > 300:
                        c = 0
            else:
                if c <= 2:
                    break
    window.close()
    return status


def adminWindow(status, users):
    sg.theme("DarkAmber")
    layout = [
        [sg.Text("Administration panel")],
        [
            [
                sg.Button(
                    f"Name: {users[i][1]}; Role: {users[i][3]}; Last login: {users[i][4]}; Last logout {users[i][5]}",
                    size=(80, 2),
                    key=f"-USER{i}-",
                )
            ]
            for i in range(len(users))
        ],
    ]
    window = sg.Window("Automatization program", layout)
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        if "-USER" in event:
            changeWindow(window, users, int(event[5:][:-1]), status)
    window.close()


def changeWindow(win, users, selected, status):
    sg.theme("DarkAmber")
    layout = [
        [sg.Text(f"Name"), sg.InputText(users[selected][1])],
        [sg.Text(f"Password"), sg.InputText()],
        [
            sg.Text(f"Role"),
            sg.Combo(
                roles,
                font=("Arial Bold", 14),
                expand_x=True,
                enable_events=True,
                readonly=True,
                key="-COMBO-",
            ),
        ],
        [sg.Button("Delete user"), sg.Button("Confirm changes")],
    ]
    window = sg.Window("Change parameters", layout)
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        if event == "Confirm changes":
            if (sanitizeInput(values[0]) != "") and (
                sanitizeInput(values[0]) != users[selected][1]
            ):
                users[selected][1] = sanitizeInput(values[0])
                changeName(status[0], users[selected][0], users[selected][1])

            if (values["-COMBO-"] != "") and (values["-COMBO-"] != users[selected][3]):
                users[selected][3] = values["-COMBO-"]
                changeRole(status[0], users[selected][0], users[selected][3])

            if sanitizeInput(values[1]) != "":
                changePassword(status[0], users[selected][0], sanitizeInput(values[1]))
            win[f"-USER{selected}-"].Update(
                f"Name: {users[selected][1]}; Role: {users[selected][3]}; Last login: {users[selected][4]}; Last logout {users[selected][5]}"
            )
            break
        if event == "Delete user":
            deleteUser(status[0], users[selected][0])
            win[f"-USER{selected}-"].Update(visible=False)
            break
    window.close()


def partsWindow(status, users, components, select=0):
    # cc = str(components[0][3]).encode(encoding="utf-16",errors="ignore")
    # print(cc)
    # print(cc.decode("utf-8", errors="ignore"))
    # print(components)
    sg.theme("DarkAmber")

    layout = [
        [sg.Text("Part selector")],
        [
            sg.Combo(
                roles,
                font=("Arial Bold", 14),
                expand_x=True,
                enable_events=True,
                readonly=True,
                key="-COMBO-",
                default_value=roles[select],
            )
        ],
        [sg.Text("Name: None"), sg.Push(), sg.Text("Art: None")],
        [sg.Text("Type: None"), sg.Push(), sg.Text("Weight: None")],
        [sg.Text("Out date: None")],
        [sg.Button("Change info")],
    ]
    window = sg.Window("Component menu", layout, element_justification="c")
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        if event == "Change info":
            changePartWindow(window)
    window.close()


def robotsWindow(status, users, components):
    sg.theme("DarkAmber")
    layout = [
        [sg.Text("Select robot")],
        [
            sg.Combo(
                roles,
                font=("Arial Bold", 14),
                expand_x=True,
                enable_events=True,
                readonly=True,
                key="-COMBO-",
            )
        ],
        [sg.Text("Name: None"), sg.Push(), sg.Text("Art: None")],
        [sg.Text("Weight: None"), sg.Push(), sg.Text("Status: None")],
        [sg.Text("Components")],
        [
            [sg.Button(f"{roles[i]}", size=(10, 1), key=f"-ITEM{i}-")]
            for i in range(len(roles))
        ],
        [sg.Button("Change info")],
    ]
    window = sg.Window("Robots menu", layout, element_justification="c")
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        if "-ITEM" in event:
            partsWindow(status, users, int(event[5:][:-1]))
        if event == "Change info":
            changeRobotWindow(window)
    window.close()


def changePartWindow(win):
    sg.theme("DarkAmber")
    iS = (20, 1)
    layout = [
        [
            sg.Text("Name: "),
            sg.InputText("", key="-NAME-", size=iS),
            sg.Push(),
            sg.Text("Art: "),
            sg.InputText("", key="-ART-", size=iS),
        ],
        [
            sg.Text("Type:  "),
            sg.InputText("", key="-TYPE-", size=iS),
            sg.Push(),
            sg.Text("Weight: "),
            sg.InputText("", key="-WEIGHT-", size=iS),
        ],
        [sg.Text("Out date: "), sg.InputText("", key="-DATE-", size=iS)],
        [sg.Button("Change info"), sg.Button("Discard changes"), sg.Button("Add new")],
    ]
    window = sg.Window("Component change", layout)
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        if event == "Change info":
            pass
        if event == "Discard changes":
            break
    window.close()


def changeRobotWindow(win):
    sg.theme("DarkAmber")
    iS = (20, 1)
    layout = [
        [
            sg.Text("Name: "),
            sg.InputText("", key="-NAME-", size=iS),
            sg.Push(),
            sg.Text("Art: "),
            sg.InputText("", key="-ART-", size=iS),
        ],
        [
            sg.Text("Weight:  "),
            sg.InputText("", key="-Weight-", size=iS),
            sg.Push(),
            sg.Text("Status: "),
            sg.InputText("", key="-Status-", size=iS),
        ],
        [sg.Text("Components: ")],
        [
            sg.Combo(
                roles,
                font=("Arial Bold", 14),
                expand_x=True,
                enable_events=True,
                readonly=True,
                key="-COMBO-",
            )
        ],
        [sg.Button("Add component"), sg.Button("Delete component")],
        [sg.Button("Change info"), sg.Button("Discard changes"), sg.Button("Add new")],
    ]
    window = sg.Window("Component change", layout)
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        if event == "Change info":
            pass
        if event == "Discard changes":
            break
    window.close()


def mainWindow(status, users):
    sg.theme("DarkAmber")
    layout = [
        [sg.Button("Admin panel", visible=(status[3] == "Administrator"))],
        [sg.Button("Component menu")],
        [sg.Button("Robot menu")],
        [sg.Button("Update components")],
    ]
    window = sg.Window(
        "Main window", layout, size=(500, 300), element_justification="c"
    )
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        if event == "Component menu":
            partsWindow(status, users, getComponents(status[0]))
        if event == "Admin panel":
            adminWindow(status, users)
        if event == "Robot menu":
            robotsWindow(status, users, getComponents(status[0]))
        if event == "Update components":
            excelToDB("./18.xlsx")
    window.close()


def main():
    initializeDatabase()
    status = authWindow()
    if status == None:
        return
    mainWindow(status, getUsers(status[0]))
    logout(status[0])
    conn.close()
    print("exit")
    sys.exit()


if __name__ == "__main__":
    main()
